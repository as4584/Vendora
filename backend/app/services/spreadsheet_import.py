"""Spreadsheet inventory import helpers.

Accepts messy reseller spreadsheets and maps common column names into the
Vendora inventory model. The router owns authentication and persistence.
"""
from __future__ import annotations

import csv
import base64
import hashlib
import io
import posixpath
import re
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any
from xml.etree import ElementTree
from urllib.parse import parse_qs, urlencode, urlparse

from fastapi import HTTPException, status


MAX_CSV_IMPORT_BYTES = 8 * 1024 * 1024
MAX_XLSX_IMPORT_BYTES = 200 * 1024 * 1024
MAX_EMBEDDED_IMAGE_BYTES = 12 * 1024 * 1024
MAX_EMBEDDED_IMAGE_DIMENSION = 640
JPEG_THUMBNAIL_QUALITY = 76

FIELD_ALIASES: dict[str, set[str]] = {
    "name": {"name", "item", "itemname", "title", "product", "productname", "description"},
    "category": {"category", "type", "department", "collection"},
    "sku": {"sku", "style", "stylecode", "stocknumber", "itemnumber", "itemid"},
    "upc": {"upc", "barcode", "ean", "gtin"},
    "size": {"size", "shoe size", "shoesize", "variant", "variantsize"},
    "color": {"color", "colour"},
    "condition": {"condition", "grade"},
    "serial_number": {"serial", "serialnumber", "imei", "watchserial"},
    "buy_price": {"buyprice", "cost", "costbasis", "purchaseprice", "paid", "wholesale"},
    "expected_sell_price": {"expectedprice", "sellprice", "listprice", "price", "askingprice", "retail"},
    "actual_sell_price": {"soldprice", "actualsellprice", "saleprice"},
    "platform": {"platform", "marketplace", "listedon", "channel"},
    "quantity": {"qty", "quantity", "stock", "count", "onhand", "inventory"},
    "vendor_name": {"vendor", "supplier", "source", "purchasedfrom"},
    "notes": {"notes", "note", "memo", "details"},
    "photo_front_url": {
        "photo",
        "photos",
        "image",
        "images",
        "imageurl",
        "image_url",
        "imagelink",
        "image link",
        "picture",
        "thumbnail",
        "productimage",
        "product image",
        "frontphoto",
        "front photo",
        "frontimage",
        "front image",
        "frontimageformula",
        "front image formula",
        "photourl",
        "photo url",
    },
    "photo_back_url": {"backphoto", "back photo", "backimage", "back image", "backimageurl", "secondaryimage"},
    "brand": {"brand", "maker", "designer"},
    "status": {"status", "state"},
}

VALID_STATUSES = {"in_stock", "listed", "sold", "shipped", "paid", "archived"}
STATUS_ALIASES = {
    "instock": "in_stock",
    "available": "in_stock",
    "active": "in_stock",
    "for sale": "listed",
    "forsale": "listed",
    "listed": "listed",
    "sold": "sold",
    "shipped": "shipped",
    "paid": "paid",
    "archived": "archived",
}
URL_RE = re.compile(r"https?://[^\s\"')<>]+", re.IGNORECASE)
SIZE_HEADER_VALUES = {"size", "sizes"}
QTY_HEADER_VALUES = {"qty", "quantity", "stock", "onhand", "qoh"}
PRODUCT_WORDS = {
    "beanie",
    "cap",
    "crewneck",
    "hoodie",
    "jacket",
    "pants",
    "shirt",
    "short",
    "shorts",
    "sweater",
    "sweatpants",
    "sweatshirt",
    "tee",
    "zip",
}
COLOR_WORDS = {
    "black",
    "blue",
    "brown",
    "camo",
    "cream",
    "green",
    "grey",
    "gray",
    "monochrome",
    "navy",
    "orange",
    "pink",
    "purple",
    "red",
    "royal",
    "vintage",
    "white",
}


@dataclass
class ParsedImportRow:
    row_number: int
    payload: dict[str, Any]
    external_id: str
    raw: dict[str, Any]
    warnings: list[str]


@dataclass(frozen=True)
class EmbeddedImage:
    row: int
    col: int
    data_url: str


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def resolve_field(header: str) -> str | None:
    normalized = normalize_header(header)
    for field, aliases in FIELD_ALIASES.items():
        if normalized in {normalize_header(alias) for alias in aliases}:
            return field
    return None


def decimal_string(value: Any) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", raw)
    if not cleaned or cleaned in {"-", "."}:
        return None
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    if amount < 0:
        return None
    return f"{amount.quantize(Decimal('0.01'))}"


def int_value(value: Any) -> int:
    raw = str(value or "").strip()
    if not raw:
        return 1
    match = re.search(r"\d+", raw)
    if not match:
        return 1
    return max(0, int(match.group(0)))


def status_value(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "in_stock"
    normalized = re.sub(r"[_\-]+", " ", raw)
    compact = normalize_header(raw)
    if raw in VALID_STATUSES:
        return raw
    return STATUS_ALIASES.get(normalized) or STATUS_ALIASES.get(compact) or "in_stock"


def looks_like_url(value: Any) -> bool:
    raw = str(value or "").strip()
    parsed = urlparse(raw)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def extract_url(value: Any) -> str | None:
    """Extract a direct URL from a plain URL, IMAGE formula, or hyperlink cell."""
    raw = str(value or "").strip()
    if not raw:
        return None
    if raw.startswith("data:image/"):
        return raw
    match = URL_RE.search(raw)
    if not match:
        return None
    return match.group(0).rstrip(".,;")


def google_sheet_csv_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.netloc not in {"docs.google.com", "www.docs.google.com"}:
        return url
    published_match = re.search(r"/spreadsheets/(?:u/\d+/)?d/e/([^/]+)", parsed.path)
    match = re.search(r"/spreadsheets/(?:u/\d+/)?d/([^/]+)", parsed.path)
    qs = parse_qs(parsed.query)
    fragment_qs = parse_qs(parsed.fragment)
    gid = (qs.get("gid") or fragment_qs.get("gid") or ["0"])[0]
    if published_match:
        query = urlencode({"single": "true", "output": "csv", "gid": gid})
        return f"https://docs.google.com/spreadsheets/d/e/{published_match.group(1)}/pub?{query}"
    if not match:
        return url
    query = urlencode({"format": "csv", "gid": gid})
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?{query}"


def google_sheet_xlsx_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.netloc not in {"docs.google.com", "www.docs.google.com"}:
        return url
    match = re.search(r"/spreadsheets/(?:u/\d+/)?d/([^/]+)", parsed.path)
    if match and match.group(1) == "e":
        return url
    if not match:
        return url
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"


def google_sheet_export_urls(url: str) -> list[str]:
    parsed = urlparse(url)
    if parsed.netloc not in {"docs.google.com", "www.docs.google.com"}:
        return [url]
    if re.search(r"/spreadsheets/(?:u/\d+/)?d/e/([^/]+)", parsed.path):
        return [google_sheet_csv_url(url)]
    if re.search(r"/spreadsheets/(?:u/\d+/)?d/([^/]+)", parsed.path):
        return [google_sheet_xlsx_url(url), google_sheet_csv_url(url)]
    return [url]


def detect_format(filename: str | None, content_type: str | None, content: bytes) -> str:
    name = (filename or "").lower()
    ctype = (content_type or "").lower()
    if name.endswith(".xlsx") or "spreadsheetml" in ctype or content.startswith(b"PK"):
        return "xlsx"
    return "csv"


def _reject_html_download(content: bytes, content_type: str | None = None) -> None:
    ctype = (content_type or "").lower()
    prefix = content[:512].lstrip().lower()
    if "text/html" in ctype or prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "The link downloaded a web page instead of spreadsheet data. "
                "Make the sheet shared as anyone-with-link read-only, or use a direct CSV/XLSX export link."
            ),
        )


def _header_score(values: list[Any]) -> int:
    fields = [resolve_field(str(value)) for value in values if str(value or "").strip()]
    unique_fields = {field for field in fields if field}
    if not unique_fields:
        return 0
    score = len(unique_fields)
    if "name" in unique_fields:
        score += 3
    if {"sku", "upc", "brand", "category"} & unique_fields:
        score += 2
    if {"buy_price", "expected_sell_price", "quantity"} & unique_fields:
        score += 2
    return score


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values, start=1):
        header = str(value or "").strip() or f"Column {idx}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        headers.append(header if count == 0 else f"{header} {count + 1}")
    return headers


def _compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _non_empty(row: list[Any], col: int) -> str:
    if col < 0 or col >= len(row):
        return ""
    return _compact_text(row[col])


def _is_matrix_size_label(value: Any) -> bool:
    label = _compact_text(value).upper()
    if not label:
        return False
    if label in {"XS", "S", "M", "L", "XL", "XXL", "XXXL", "OS", "O/S", "ONE SIZE"}:
        return True
    return bool(re.fullmatch(r"\d{1,2}(\.\d)?", label))


def _size_quantity_pairs(row: list[Any]) -> list[tuple[int, int]]:
    pairs: list[tuple[int, int]] = []
    for idx, value in enumerate(row):
        if normalize_header(value) not in SIZE_HEADER_VALUES:
            continue
        for qty_col in range(idx + 1, min(idx + 4, len(row))):
            if normalize_header(row[qty_col]) in QTY_HEADER_VALUES:
                pairs.append((idx, qty_col))
                break
    return pairs


def _matrix_quantity(value: Any) -> int:
    raw = str(value or "").strip()
    if not raw:
        return 0
    match = re.search(r"\d+", raw)
    if not match:
        return 0
    return max(0, int(match.group(0)))


def _find_matrix_product_name(
    table_rows: list[list[Any]],
    header_idx: int,
    pairs: list[tuple[int, int]],
    pair_idx: int,
) -> tuple[str, int, int] | tuple[None, None, None]:
    size_col, qty_col = pairs[pair_idx]
    left_bound = pairs[pair_idx - 1][1] + 1 if pair_idx > 0 else 0
    right_bound = pairs[pair_idx + 1][0] - 1 if pair_idx + 1 < len(pairs) else qty_col + 2
    right_bound = max(left_bound, min(right_bound, max((len(row) for row in table_rows), default=0) - 1))

    preferred_cols = [size_col - 1, size_col, qty_col - 1, qty_col]
    preferred_cols.extend(range(left_bound, right_bound + 1))
    ordered_cols = list(dict.fromkeys(col for col in preferred_cols if left_bound <= col <= right_bound))

    for row_idx in range(header_idx - 1, max(-1, header_idx - 8), -1):
        row = table_rows[row_idx]
        for col in ordered_cols:
            candidate = _non_empty(row, col)
            normalized = normalize_header(candidate)
            if not candidate or normalized in SIZE_HEADER_VALUES | QTY_HEADER_VALUES:
                continue
            if _is_matrix_size_label(candidate) or candidate.isdigit():
                continue
            return candidate, row_idx + 1, col + 1
    return None, None, None


def _collect_matrix_variants(
    table_rows: list[list[Any]],
    header_idx: int,
    pairs: list[tuple[int, int]],
) -> dict[int, list[dict[str, int | str]]]:
    variants_by_pair: dict[int, list[dict[str, int | str]]] = {idx: [] for idx in range(len(pairs))}
    relevant_cols = {col for pair in pairs for col in pair}

    for row in table_rows[header_idx + 1:]:
        if _size_quantity_pairs(row):
            break
        if not any(_non_empty(row, col) for col in relevant_cols):
            break

        for pair_idx, (size_col, qty_col) in enumerate(pairs):
            size = _non_empty(row, size_col)
            if not size or normalize_header(size) in SIZE_HEADER_VALUES:
                continue
            if not _is_matrix_size_label(size):
                continue
            variants_by_pair[pair_idx].append({
                "size": size,
                "quantity": _matrix_quantity(_non_empty(row, qty_col)),
            })

    return variants_by_pair


def _infer_category(name: str) -> str | None:
    normalized = normalize_header(name)
    if "sweatshort" in normalized or "short" in normalized:
        return "Shorts"
    if "sweatpant" in normalized:
        return "Sweatpants"
    if "hoodie" in normalized or "zipup" in normalized:
        return "Hoodie"
    if "beanie" in normalized or "cap" in normalized:
        return "Headwear"
    if "shirt" in normalized or normalized.endswith("tee"):
        return "Shirt"
    return None


def _infer_color(raw_name: str) -> str | None:
    lines = [line.strip() for line in str(raw_name or "").splitlines() if line.strip()]
    if not lines:
        return None

    tail_words = re.findall(r"[A-Za-z]+", lines[-1].lower())
    if not tail_words:
        return None

    if not any(word in COLOR_WORDS for word in tail_words):
        return None

    if not any(word in PRODUCT_WORDS for word in tail_words):
        return lines[-1][:50]

    color_tail: list[str] = []
    for word in reversed(tail_words):
        if word not in COLOR_WORDS:
            break
        color_tail.append(word)
    if not color_tail:
        return None
    return " ".join(reversed(color_tail)).title()[:50]


def _warehouse_matrix_rows_to_dicts(
    table_rows: list[list[Any]],
    embedded_images: list[EmbeddedImage],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for header_idx, row in enumerate(table_rows):
        pairs = _size_quantity_pairs(row)
        if not pairs:
            continue

        variants_by_pair = _collect_matrix_variants(table_rows, header_idx, pairs)
        for pair_idx, variants in variants_by_pair.items():
            if not variants:
                continue
            product_name, product_row_number, product_col_number = _find_matrix_product_name(
                table_rows,
                header_idx,
                pairs,
                pair_idx,
            )
            if not product_name or product_row_number is None or product_col_number is None:
                continue

            name = _compact_text(product_name)
            total_quantity = sum(int(variant["quantity"]) for variant in variants)
            size_col, qty_col = pairs[pair_idx]
            left_bound = pairs[pair_idx - 1][1] + 1 if pair_idx > 0 else 0
            right_bound = pairs[pair_idx + 1][0] - 1 if pair_idx + 1 < len(pairs) else qty_col + 2
            right_bound = max(left_bound, min(right_bound, max((len(row) for row in table_rows), default=0) - 1))
            parsed_row: dict[str, Any] = {
                "Product Name": name,
                "Qty": str(total_quantity),
                "__variants": variants,
                "__layout": "size_quantity_matrix",
                "__row_number": product_row_number,
            }
            image_url = _closest_matrix_image(
                embedded_images,
                product_row=product_row_number,
                product_col=product_col_number,
                header_row=header_idx + 1,
                left_bound=left_bound,
                right_bound=right_bound,
            )
            if image_url:
                parsed_row["Image URL"] = image_url
            category = _infer_category(name)
            if category:
                parsed_row["Category"] = category
            color = _infer_color(str(product_name))
            if color:
                parsed_row["Color"] = color
            rows.append(parsed_row)

    return rows


def _image_data_url(image_bytes: bytes, path: str) -> str | None:
    if len(image_bytes) > MAX_EMBEDDED_IMAGE_BYTES:
        return None

    try:
        from PIL import Image
    except ImportError:
        Image = None

    if Image is not None:
        try:
            with Image.open(io.BytesIO(image_bytes)) as image:
                image.thumbnail(
                    (MAX_EMBEDDED_IMAGE_DIMENSION, MAX_EMBEDDED_IMAGE_DIMENSION),
                    Image.Resampling.LANCZOS,
                )
                if image.mode not in {"RGB", "L"}:
                    image = image.convert("RGB")
                output = io.BytesIO()
                image.save(output, format="JPEG", quality=JPEG_THUMBNAIL_QUALITY)
                encoded = base64.b64encode(output.getvalue()).decode("ascii")
                return f"data:image/jpeg;base64,{encoded}"
        except Exception:
            pass

    extension = path.rsplit(".", 1)[-1].lower()
    mime = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "webp": "image/webp",
    }.get(extension)
    if not mime:
        return None
    encoded = base64.b64encode(image_bytes).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _resolve_xlsx_relationship_path(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_path), target))


def _xlsx_first_sheet_images(content: bytes) -> list[EmbeddedImage]:
    namespaces = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    }

    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        return []

    try:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        sheet_rels_xml = archive.read("xl/worksheets/_rels/sheet1.xml.rels")
    except KeyError:
        return []

    sheet_root = ElementTree.fromstring(sheet_xml)
    drawing = sheet_root.find("main:drawing", namespaces)
    if drawing is None:
        return []
    drawing_rel_id = drawing.attrib.get(f"{{{namespaces['r']}}}id")
    if not drawing_rel_id:
        return []

    sheet_rels = ElementTree.fromstring(sheet_rels_xml)
    drawing_target = None
    for relationship in sheet_rels.findall("rel:Relationship", namespaces):
        if relationship.attrib.get("Id") == drawing_rel_id:
            drawing_target = relationship.attrib.get("Target")
            break
    if not drawing_target:
        return []

    drawing_path = _resolve_xlsx_relationship_path("xl/worksheets/sheet1.xml", drawing_target)
    drawing_rels_path = posixpath.join(
        posixpath.dirname(drawing_path),
        "_rels",
        f"{posixpath.basename(drawing_path)}.rels",
    )

    try:
        drawing_root = ElementTree.fromstring(archive.read(drawing_path))
        drawing_rels = ElementTree.fromstring(archive.read(drawing_rels_path))
    except KeyError:
        return []

    media_targets = {
        relationship.attrib.get("Id"): relationship.attrib.get("Target")
        for relationship in drawing_rels.findall("rel:Relationship", namespaces)
    }

    images: list[EmbeddedImage] = []
    for anchor_name in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in drawing_root.findall(f"xdr:{anchor_name}", namespaces):
            marker = anchor.find("xdr:from", namespaces)
            blip = anchor.find(".//a:blip", namespaces)
            if marker is None or blip is None:
                continue
            embed_id = blip.attrib.get(f"{{{namespaces['r']}}}embed")
            media_target = media_targets.get(embed_id)
            if not media_target:
                continue

            media_path = _resolve_xlsx_relationship_path(drawing_path, media_target)
            try:
                data_url = _image_data_url(archive.read(media_path), media_path)
            except KeyError:
                data_url = None
            if not data_url:
                continue

            row = int(marker.findtext("xdr:row", default="0", namespaces=namespaces)) + 1
            col = int(marker.findtext("xdr:col", default="0", namespaces=namespaces)) + 1
            images.append(EmbeddedImage(row=row, col=col, data_url=data_url))
    return images


def _closest_matrix_image(
    images: list[EmbeddedImage],
    *,
    product_row: int,
    product_col: int,
    header_row: int,
    left_bound: int,
    right_bound: int,
) -> str | None:
    candidates: list[tuple[int, EmbeddedImage]] = []
    for image in images:
        if not (product_row <= image.row <= header_row):
            continue
        if not (left_bound + 1 <= image.col <= right_bound + 1):
            continue
        score = abs(image.row - product_row) * 10 + abs(image.col - product_col)
        candidates.append((score, image))
    if not candidates:
        return None
    return min(candidates, key=lambda item: item[0])[1].data_url


def _table_rows_to_dicts(
    table_rows: list[list[Any]],
    embedded_images: list[EmbeddedImage] | None = None,
) -> list[dict[str, Any]]:
    matrix_rows = _warehouse_matrix_rows_to_dicts(table_rows, embedded_images or [])
    if matrix_rows:
        return matrix_rows

    candidates = [
        (idx, _header_score(row))
        for idx, row in enumerate(table_rows[:25])
        if any(str(value or "").strip() for value in row)
    ]
    if not candidates:
        return []

    header_idx, score = max(candidates, key=lambda item: item[1])
    if score < 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Could not find an inventory header row. Add headers like Product Name, SKU, "
                "Brand, Cost, List Price, Qty, or Image URL."
            ),
        )

    headers = _unique_headers(table_rows[header_idx])
    rows: list[dict[str, Any]] = []
    for row in table_rows[header_idx + 1:]:
        if not any(value not in (None, "") for value in row):
            continue
        rows.append({
            headers[idx]: value
            for idx, value in enumerate(row)
            if idx < len(headers) and value not in (None, "")
        })
    return rows


def rows_from_bytes(content: bytes, file_format: str, content_type: str | None = None) -> list[dict[str, Any]]:
    max_bytes = MAX_XLSX_IMPORT_BYTES if file_format == "xlsx" else MAX_CSV_IMPORT_BYTES
    max_mb = max_bytes // (1024 * 1024)
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Spreadsheet is too large. Import files must be {max_mb} MB or less.",
        )
    _reject_html_download(content, content_type)

    if file_format == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="XLSX import support is not installed on the server.",
            ) from exc

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            table_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        embedded_images = _xlsx_first_sheet_images(content)
        return _table_rows_to_dicts(table_rows, embedded_images=embedded_images)

    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    except csv.Error:
        dialect = csv.excel
    if getattr(dialect, "delimiter", ",") in {"\r", "\n"}:
        dialect = csv.excel
    return _table_rows_to_dicts([list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)])


def parse_inventory_rows(rows: list[dict[str, Any]]) -> list[ParsedImportRow]:
    parsed_rows: list[ParsedImportRow] = []

    for idx, raw in enumerate(rows, start=2):
        row_number = idx
        raw_row_number = raw.get("__row_number")
        if isinstance(raw_row_number, int):
            row_number = raw_row_number
        mapped: dict[str, Any] = {}
        custom_attributes: dict[str, Any] = {}
        warnings: list[str] = []

        for header, value in raw.items():
            if header == "__row_number":
                continue
            if value is None or str(value).strip() == "":
                continue
            if str(header).startswith("__"):
                attribute_name = {
                    "__variants": "variants",
                    "__layout": "import_layout",
                }.get(str(header))
                if attribute_name:
                    custom_attributes[attribute_name] = value
                continue
            field = resolve_field(header)
            if not field:
                custom_attributes[normalize_header(header) or str(header)] = value
                continue
            mapped[field] = str(value).strip()

        if not mapped.get("name"):
            brand = mapped.get("brand")
            sku = mapped.get("sku") or mapped.get("upc")
            category = mapped.get("category")
            mapped["name"] = " ".join(part for part in [brand, category, sku] if part).strip()

        if not mapped.get("name"):
            warnings.append("Skipped row without item name or usable identifier.")
            parsed_rows.append(ParsedImportRow(row_number, {}, "", raw, warnings))
            continue

        payload: dict[str, Any] = {
            "name": mapped["name"][:255],
            "category": mapped.get("category"),
            "sku": mapped.get("sku"),
            "upc": mapped.get("upc"),
            "size": mapped.get("size"),
            "color": mapped.get("color"),
            "condition": mapped.get("condition"),
            "serial_number": mapped.get("serial_number"),
            "buy_price": decimal_string(mapped.get("buy_price")),
            "expected_sell_price": decimal_string(mapped.get("expected_sell_price")),
            "actual_sell_price": decimal_string(mapped.get("actual_sell_price")),
            "platform": mapped.get("platform"),
            "quantity": int_value(mapped.get("quantity")),
            "vendor_name": mapped.get("vendor_name"),
            "notes": mapped.get("notes"),
            "status": status_value(mapped.get("status")),
        }

        if mapped.get("brand"):
            custom_attributes["brand"] = mapped["brand"]
        photo_front_url = extract_url(mapped.get("photo_front_url"))
        if photo_front_url:
            payload["photo_front_url"] = photo_front_url
        elif mapped.get("photo_front_url"):
            warnings.append("Ignored photo column because it was not a direct http(s) URL.")
        photo_back_url = extract_url(mapped.get("photo_back_url"))
        if photo_back_url:
            payload["photo_back_url"] = photo_back_url

        custom_attributes["import_raw"] = {
            str(k): v for k, v in raw.items() if not str(k).startswith("__") and v not in (None, "")
        }
        payload["custom_attributes"] = custom_attributes
        payload = {k: v for k, v in payload.items() if v not in (None, "")}

        fingerprint_base = "|".join(
            str(payload.get(field, "")) for field in ("sku", "upc", "name", "size", "color")
        )
        external_id = hashlib.sha256(fingerprint_base.encode("utf-8")).hexdigest()[:32]
        parsed_rows.append(ParsedImportRow(row_number, payload, external_id, raw, warnings))

    return parsed_rows
