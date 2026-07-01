"""XLSX inventory export with **embedded** item photos.

Photos are stored as base64 data URLs, which `=IMAGE()` in Sheets/Excel cannot
render (it only accepts public http(s) URLs) — so the CSV export just showed a
giant base64 string. This export decodes each photo and embeds the real image
into the cell, producing a clean, image-rich spreadsheet that works everywhere.
"""
from __future__ import annotations

import base64
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy.orm import Session

from app.models.inventory import InventoryItem

# Text columns (the base64 URL + formula columns are intentionally dropped —
# replaced by two columns of real embedded thumbnails).
_TEXT_COLUMNS = [
    "id", "name", "category", "sku", "upc", "size", "color", "condition",
    "quantity", "buy_price", "expected_sell_price", "actual_sell_price",
    "status", "platform", "vendor_name", "notes", "source", "external_id",
    "created_at", "updated_at",
]
_PHOTO_COLUMNS = ["Front Photo", "Back Photo"]

_THUMB_PX = 96
_HEADER_FILL = PatternFill("solid", fgColor="F26722")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _decode_image(url: str | None) -> bytes | None:
    """Return PNG bytes for a base64 data URL (or None). Resized to a thumbnail."""
    if not url or not isinstance(url, str) or not url.startswith("data:"):
        return None
    try:
        _, b64 = url.split(",", 1)
        raw = base64.b64decode(b64)
        im = PILImage.open(io.BytesIO(raw))
        im.thumbnail((_THUMB_PX, _THUMB_PX))
        if im.mode not in ("RGB", "RGBA"):
            im = im.convert("RGBA")
        out = io.BytesIO()
        im.save(out, format="PNG")
        out.seek(0)
        return out.getvalue()
    except Exception:
        return None


def _resolved_photo(item: InventoryItem, key: str) -> str:
    direct = getattr(item, f"{key}_url", None)
    if direct:
        return direct
    attrs = item.custom_attributes or {}
    return attrs.get(key) or ""


def export_inventory_xlsx(db: Session, user_id) -> bytes:
    items = (
        db.query(InventoryItem)
        .filter(InventoryItem.user_id == user_id, InventoryItem.deleted_at.is_(None))
        .order_by(InventoryItem.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = _TEXT_COLUMNS + _PHOTO_COLUMNS
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(vertical="center")

    # widen photo columns so thumbnails fit
    front_col = len(_TEXT_COLUMNS) + 1
    back_col = len(_TEXT_COLUMNS) + 2
    for col in (front_col, back_col):
        ws.column_dimensions[get_column_letter(col)].width = _THUMB_PX / 7.0

    for r, item in enumerate(items, start=2):
        values = [
            str(item.id), item.name, item.category or "", item.sku or "", item.upc or "",
            item.size or "", item.color or "", item.condition or "", item.quantity,
            str(item.buy_price) if item.buy_price is not None else "",
            str(item.expected_sell_price) if item.expected_sell_price is not None else "",
            str(item.actual_sell_price) if item.actual_sell_price is not None else "",
            item.status, item.platform or "", item.vendor_name or "", item.notes or "",
            item.source or "", item.external_id or "",
            item.created_at.isoformat() if item.created_at else "",
            item.updated_at.isoformat() if item.updated_at else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=r, column=col_idx, value=val)

        has_image = False
        for key, col in (("photo_front", front_col), ("photo_back", back_col)):
            png = _decode_image(_resolved_photo(item, key))
            if png:
                xlimg = XLImage(io.BytesIO(png))
                xlimg.width, xlimg.height = _THUMB_PX, _THUMB_PX
                ws.add_image(xlimg, f"{get_column_letter(col)}{r}")
                has_image = True
        if has_image:
            ws.row_dimensions[r].height = _THUMB_PX * 0.75  # px → points

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
