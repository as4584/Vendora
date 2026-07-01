"""Focused branch coverage for spreadsheet import's defensive helpers."""
from __future__ import annotations

import io
import builtins
import sys
import zipfile

import pytest
from fastapi import HTTPException
from PIL import Image
from openpyxl import Workbook
import openpyxl

from app.services import spreadsheet_import as subject


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (None, None),
        ("$1,234.5", "1234.50"),
        ("-", None),
        ("not a price", None),
        ("-2", None),
    ],
)
def test_decimal_string_defensive_inputs(value, expected):
    assert subject.decimal_string(value) == expected


def test_decimal_string_rejects_malformed_numeric_text():
    assert subject.decimal_string("1.2.3") is None


@pytest.mark.parametrize(
    ("value", "expected"),
    [(None, 1), ("many", 1), ("qty: 7", 7), ("0", 0)],
)
def test_int_value_defensive_inputs(value, expected):
    assert subject.int_value(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [(None, "in_stock"), ("in_stock", "in_stock"), ("FOR-SALE", "listed"), ("unknown", "in_stock")],
)
def test_status_value_normalizes_aliases(value, expected):
    assert subject.status_value(value) == expected


def test_url_extractors_accept_supported_forms_only():
    data_url = "data:image/png;base64,abc"
    assert subject.extract_url(data_url) == data_url
    assert subject.extract_url('=IMAGE("https://example.test/a.png"),') == "https://example.test/a.png"
    assert subject.extract_url("not a link") is None
    assert subject.looks_like_url("https://example.test/a")
    assert not subject.looks_like_url("file:///tmp/a")


def test_google_sheet_export_variants():
    plain = "https://example.test/sheet.csv"
    published = "https://docs.google.com/spreadsheets/d/e/PUB-ID/pubhtml#gid=42"
    editable = "https://docs.google.com/spreadsheets/u/1/d/SHEET-ID/edit?gid=9"
    malformed = "https://docs.google.com/not-a-sheet"

    assert subject.google_sheet_csv_url(plain) == plain
    assert "d/e/PUB-ID/pub?" in subject.google_sheet_csv_url(published)
    assert "gid=42" in subject.google_sheet_csv_url(published)
    assert "format=csv" in subject.google_sheet_csv_url(editable)
    assert subject.google_sheet_csv_url(malformed) == malformed
    assert subject.google_sheet_xlsx_url(plain) == plain
    assert subject.google_sheet_xlsx_url(published) == published
    assert subject.google_sheet_xlsx_url(malformed) == malformed
    assert subject.google_sheet_export_urls(plain) == [plain]
    assert len(subject.google_sheet_export_urls(editable)) == 2
    assert len(subject.google_sheet_export_urls(published)) == 1
    assert subject.google_sheet_export_urls(malformed) == [malformed]


def test_google_sheet_candidate_urls_deduplicate_and_limit():
    url = "https://docs.google.com/spreadsheets/d/SHEET-ID/edit"
    html = 'gid=12 gid=12 "sheetId":34 gid=56'
    candidates = subject.google_sheet_candidate_csv_urls(url, html, limit=2)
    assert len(candidates) == 2
    assert "gid=12" in candidates[0]
    assert "gid=56" in candidates[1]
    assert subject.google_sheet_candidate_csv_urls("https://example.test/a", html) == []
    assert subject.google_sheet_candidate_csv_urls(url, "no sheet ids here") == []


def test_format_detection_and_html_rejection():
    assert subject.detect_format("items.xlsx", None, b"data") == "xlsx"
    assert subject.detect_format(None, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"data") == "xlsx"
    assert subject.detect_format("items.csv", "text/csv", b"PKzip") == "xlsx"
    assert subject.detect_format("items.csv", "text/csv", b"a,b") == "csv"
    for content, content_type in ((b"<html>no</html>", None), (b" <!doctype html>", None), (b"x", "text/html")):
        with pytest.raises(HTTPException) as exc:
            subject._reject_html_download(content, content_type)
        assert exc.value.status_code == 400


def test_header_and_matrix_primitives():
    assert subject._header_score(["nothing"]) == 0
    assert subject._header_score(["Product Name", "SKU", "Cost"]) == 10
    assert subject._header_score(["SKU", "Cost"]) == 6
    assert subject._unique_headers(["Name", "Name", None]) == ["Name", "Name 2", "Column 3"]
    assert subject._non_empty([" a  b "], 0) == "a b"
    assert subject._non_empty([], -1) == ""
    assert subject._non_empty([], 2) == ""
    assert subject._is_matrix_size_label("O/S")
    assert subject._is_matrix_size_label("10.5")
    assert not subject._is_matrix_size_label("purple")
    assert not subject._is_matrix_size_label("")
    assert subject._size_quantity_pairs(["Size", "x", "Qty"]) == [(0, 2)]
    assert subject._size_quantity_pairs(["Size", "x", "x", "x", "Qty"]) == []
    assert subject._matrix_quantity(None) == 0
    assert subject._matrix_quantity("many") == 0
    assert subject._matrix_quantity("qty 4") == 4


def test_matrix_name_and_variant_collection_edges():
    rows = [["Blue Hoodie", "", ""], ["", "Size", "Qty"], ["", "S", "2"], ["", "M", "many"], ["", "Size", "Qty"]]
    name, row, col = subject._find_matrix_product_name(rows, 1, [(1, 2)], 0)
    assert (name, row, col) == ("Blue Hoodie", 1, 1)
    assert subject._collect_matrix_variants(rows, 1, [(1, 2)]) == {
        0: [{"size": "S", "quantity": 2}, {"size": "M", "quantity": 0}]
    }
    assert subject._find_matrix_product_name([["Size", "Qty"]], 0, [(0, 1)], 0) == (None, None, None)
    assert subject._find_matrix_product_name([["10", ""], ["Size", "Qty"]], 1, [(0, 1)], 0) == (None, None, None)
    assert subject._collect_matrix_variants([["Size", "Qty"], ["", "1"]], 0, [(0, 1)]) == {0: []}
    assert subject._collect_matrix_variants([["Size", "Qty"], ["Purple", "1"]], 0, [(0, 1)]) == {0: []}
    assert subject._collect_matrix_variants([["Size", "Qty"], ["", ""]], 0, [(0, 1)]) == {0: []}


@pytest.mark.parametrize(
    ("name", "expected"),
    [("Running shorts", "Shorts"), ("Cozy sweatpants", "Sweatpants"), ("Zip-up", "Hoodie"), ("Wool cap", "Headwear"), ("Logo tee", "Shirt"), ("Watch", None)],
)
def test_category_inference(name, expected):
    assert subject._infer_category(name) == expected


def test_color_and_size_inference_edges():
    assert subject._infer_color("") is None
    assert subject._infer_color("123") is None
    assert subject._infer_color("Product\nSpecial Edition") is None
    assert subject._infer_color("Product\nNavy") == "Navy"
    assert subject._infer_color("Hoodie Black") == "Black"
    assert subject._infer_color("Red Hoodie") is None
    assert subject._display_size_label("medium") == "M"
    assert subject._display_size_label("xl") == "XL"


def test_image_encoding_fallbacks_and_size_limit(monkeypatch):
    assert subject._image_data_url(b"x" * (subject.MAX_EMBEDDED_IMAGE_BYTES + 1), "a.png") is None
    assert subject._image_data_url(b"not-an-image", "a.unknown") is None
    assert subject._image_data_url(b"not-an-image", "a.png").startswith("data:image/png;base64,")
    png = io.BytesIO()
    Image.new("RGBA", (2, 2), (255, 0, 0, 128)).save(png, format="PNG")
    assert subject._image_data_url(png.getvalue(), "a.png").startswith("data:image/jpeg;base64,")

    real_import = builtins.__import__

    def without_pillow(name, *args, **kwargs):
        if name == "PIL":
            raise ImportError("Pillow unavailable")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", without_pillow)
    assert subject._image_data_url(b"raw", "a.png").startswith("data:image/png;base64,")


def test_relationship_and_closest_image_helpers():
    assert subject._resolve_xlsx_relationship_path("xl/a.xml", "/xl/media/a.png") == "xl/media/a.png"
    assert subject._resolve_xlsx_relationship_path("xl/a.xml", "media/a.png") == "xl/media/a.png"
    images = [
        subject.EmbeddedImage(2, 2, "far"),
        subject.EmbeddedImage(1, 1, "near"),
        subject.EmbeddedImage(20, 20, "outside"),
    ]
    assert subject._closest_matrix_image(images, product_row=1, product_col=1, header_row=3, left_bound=0, right_bound=2) == "near"
    assert subject._closest_matrix_image(images, product_row=5, product_col=5, header_row=6, left_bound=4, right_bound=5) is None
    assert subject._closest_matrix_image(
        [subject.EmbeddedImage(2, 20, "wrong-column")],
        product_row=1, product_col=1, header_row=3, left_bound=0, right_bound=2,
    ) is None


def test_table_and_csv_parsing_defensive_paths(monkeypatch):
    assert subject._table_rows_to_dicts([]) == []
    with pytest.raises(HTTPException, match="Could not find"):
        subject._table_rows_to_dicts([["random", "values"]])
    rows = subject._table_rows_to_dicts([["Product Name", "SKU", "SKU", None], [None, "", "", ""], ["Hat", "A", "B", ""]])
    assert rows == [{"Product Name": "Hat", "SKU": "A", "SKU 2": "B"}]
    with pytest.raises(HTTPException) as exc:
        subject.rows_from_bytes(b"x" * (subject.MAX_CSV_IMPORT_BYTES + 1), "csv")
    assert exc.value.status_code == 413
    with pytest.raises(HTTPException, match="web page"):
        subject.rows_from_bytes(b"<html>oops", "csv")
    assert subject.rows_from_bytes(b"Product Name,SKU\nHat,A\n", "csv") == [{"Product Name": "Hat", "SKU": "A"}]
    monkeypatch.setattr(subject.csv.Sniffer, "sniff", lambda *args, **kwargs: (_ for _ in ()).throw(subject.csv.Error()))
    assert subject.rows_from_bytes(b"Product Name,SKU\nHat,A\n", "csv") == [{"Product Name": "Hat", "SKU": "A"}]


def test_horizontal_and_warehouse_layout_edge_rows():
    horizontal = [
        ["ESSENTIALS", "INVENTORY", ""],
        ["ITEM NAME", "XS", "TOTAL"],
        ["", "1", "1"],
        ["Zero", "0", "0"],
        ["Hoodie Navy", "0", "3"],
        ["ITEM NAME", "XS", "TOTAL"],
    ]
    rows = subject._horizontal_size_table_rows_to_dicts(horizontal)
    assert rows == [{
        "Product Name": "Hoodie Navy",
        "Qty": "3",
        "__variants": [{"size": "XS", "quantity": 0}],
        "__layout": "horizontal_size_table",
        "__row_number": 5,
        "Category": "ESSENTIALS",
        "Color": "Navy",
    }]
    assert subject._horizontal_size_table_rows_to_dicts_should_stop(["INVENTORY"], 0)
    no_category = subject._horizontal_size_table_rows_to_dicts([
        ["ITEM NAME", "S", "TOTAL"],
        ["Plain Product", "1", "1"],
    ])
    assert "Category" not in no_category[0]
    assert subject._warehouse_matrix_rows_to_dicts([["Size", "Qty"], ["Purple", "1"]], []) == []
    assert subject._warehouse_matrix_rows_to_dicts([["Size", "Qty"], ["S", "1"]], []) == []
    plain_matrix = subject._warehouse_matrix_rows_to_dicts([
        ["Watch"],
        ["Size", "Qty"],
        ["S", "1"],
    ], [])
    assert plain_matrix[0]["Product Name"] == "Watch"
    assert "Category" not in plain_matrix[0] and "Color" not in plain_matrix[0]


def _zip_bytes(files):
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        for path, content in files.items():
            archive.writestr(path, content)
    return output.getvalue()


def test_xlsx_image_reader_rejects_incomplete_archives():
    assert subject._xlsx_first_sheet_images(b"not-a-zip") == []
    assert subject._xlsx_first_sheet_images(_zip_bytes({"other": "x"})) == []
    sheet_no_drawing = '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" />'
    rels = '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships" />'
    base = {
        "xl/worksheets/sheet1.xml": sheet_no_drawing,
        "xl/worksheets/_rels/sheet1.xml.rels": rels,
    }
    assert subject._xlsx_first_sheet_images(_zip_bytes(base)) == []

    sheet_missing_id = '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><drawing /></worksheet>'
    assert subject._xlsx_first_sheet_images(_zip_bytes({**base, "xl/worksheets/sheet1.xml": sheet_missing_id})) == []

    sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<drawing r:id="rId1" /></worksheet>'
    )
    assert subject._xlsx_first_sheet_images(_zip_bytes({**base, "xl/worksheets/sheet1.xml": sheet})) == []
    rels_with_target = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Target="../drawings/drawing1.xml" /></Relationships>'
    )
    assert subject._xlsx_first_sheet_images(_zip_bytes({
        **base,
        "xl/worksheets/sheet1.xml": sheet,
        "xl/worksheets/_rels/sheet1.xml.rels": rels_with_target,
    })) == []

    sheet_rid2 = sheet.replace('r:id="rId1"', 'r:id="rId2"')
    two_rels = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Target="../drawings/ignored.xml" />'
        '<Relationship Id="rId2" Target="../drawings/drawing2.xml" /></Relationships>'
    )
    assert subject._xlsx_first_sheet_images(_zip_bytes({
        **base,
        "xl/worksheets/sheet1.xml": sheet_rid2,
        "xl/worksheets/_rels/sheet1.xml.rels": two_rels,
    })) == []


def test_xlsx_image_reader_skips_malformed_anchors_and_missing_media():
    sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<drawing r:id="rId1" /></worksheet>'
    )
    sheet_rels = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Target="../drawings/drawing1.xml" /></Relationships>'
    )
    drawing = (
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<xdr:oneCellAnchor><a:blip r:embed="rIdImage" /></xdr:oneCellAnchor>'
        '<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from><a:blip r:embed="unknown" /></xdr:oneCellAnchor>'
        '<xdr:twoCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from><a:blip r:embed="rIdImage" /></xdr:twoCellAnchor>'
        '</xdr:wsDr>'
    )
    drawing_rels = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdImage" Target="../media/missing.png" /></Relationships>'
    )
    content = _zip_bytes({
        "xl/worksheets/sheet1.xml": sheet,
        "xl/worksheets/_rels/sheet1.xml.rels": sheet_rels,
        "xl/drawings/drawing1.xml": drawing,
        "xl/drawings/_rels/drawing1.xml.rels": drawing_rels,
    })
    assert subject._xlsx_first_sheet_images(content) == []


def test_xlsx_import_missing_dependency_and_invalid_sheet(monkeypatch):
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    with pytest.raises(HTTPException) as exc:
        subject.rows_from_bytes(b"PK placeholder", "xlsx")
    assert exc.value.status_code == 500


def test_xlsx_import_raises_last_header_error_after_all_sheets():
    workbook = Workbook()
    workbook.active.append(["random", "values"])
    workbook.create_sheet("Second").append(["still", "not inventory"])
    output = io.BytesIO()
    workbook.save(output)
    with pytest.raises(HTTPException, match="Could not find"):
        subject.rows_from_bytes(output.getvalue(), "xlsx")


def test_xlsx_import_empty_workbook_returns_no_rows(monkeypatch):
    fake = type("Workbook", (), {"worksheets": [], "close": lambda self: None})()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: fake)
    assert subject.rows_from_bytes(b"PK placeholder", "xlsx") == []


def test_parse_inventory_rows_handles_skips_synthesis_and_review_flags():
    rows = [
        {"__row_number": 8, "Unknown": "", "__layout": "matrix", "__ignored": "value"},
        {"Brand": "Acme", "Category": "Hat", "SKU": "A-1", "Photo": "not a url", "Odd Field": "kept", "Status": "for sale"},
        {"Product Name": "Complete", "Size": "M", "Cost": "$2", "Front Photo": "https://example.test/front.jpg", "Back Photo": "https://example.test/back.jpg", "__variants": [{"size": "M", "quantity": 1}]},
    ]
    parsed = subject.parse_inventory_rows(rows)
    assert parsed[0].row_number == 8
    assert parsed[0].payload == {}
    assert parsed[0].external_id == ""
    assert "without item name" in parsed[0].warnings[0]
    assert parsed[1].payload["name"] == "Acme Hat A-1"
    assert parsed[1].payload["status"] == "listed"
    assert parsed[1].payload["custom_attributes"]["oddfield"] == "kept"
    assert parsed[1].payload["custom_attributes"]["import_review"] == {
        "missing_price": True,
        "missing_size": True,
        "missing_photo": True,
    }
    assert any("Ignored photo" in warning for warning in parsed[1].warnings)
    assert parsed[2].payload["photo_front_url"].startswith("https://")
    assert parsed[2].payload["photo_back_url"].startswith("https://")
    assert len(parsed[2].external_id) == 32
