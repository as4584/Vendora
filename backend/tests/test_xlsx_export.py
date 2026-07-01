"""XLSX inventory export — item photos embedded as real images (not base64 text)."""
import base64
import io

from openpyxl import load_workbook
from PIL import Image as PILImage

from app.models.inventory import InventoryItem
from app.services.xlsx_export import _decode_image, export_inventory_xlsx


def _data_url(color=(242, 103, 34)) -> str:
    im = PILImage.new("RGB", (150, 150), color)
    b = io.BytesIO()
    im.save(b, "JPEG")
    return "data:image/jpeg;base64," + base64.b64encode(b.getvalue()).decode()


def test_decode_image_handles_data_url_and_junk():
    assert _decode_image(_data_url()) is not None      # valid base64 image -> PNG bytes
    assert _decode_image("") is None
    assert _decode_image("https://x/y.jpg") is None     # not a data URL -> skipped
    assert _decode_image("data:image/jpeg;base64,not-base64!!") is None  # bad payload -> None (no crash)


def test_xlsx_embeds_real_image_not_base64_text(db, test_user):
    url = _data_url()
    item = InventoryItem(user_id=test_user.id, name="Epoxy Kit", status="in_stock",
                         quantity=2, photo_front_url=url)
    db.add(item)
    db.commit()

    xlsx = export_inventory_xlsx(db, test_user.id)
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active

    # a real image is embedded (the whole point)
    assert len(ws._images) == 1
    # and the giant base64 string is NOT dumped into any cell
    all_text = "".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "base64" not in all_text
    assert "Epoxy Kit" in all_text          # normal text columns still present
    assert ws.cell(row=1, column=ws.max_column).value == "Back Photo"  # photo columns exist


def test_xlsx_ok_when_no_photo(db, test_user):
    db.add(InventoryItem(user_id=test_user.id, name="No Photo", status="in_stock", quantity=1))
    db.commit()
    xlsx = export_inventory_xlsx(db, test_user.id)
    ws = load_workbook(io.BytesIO(xlsx)).active
    assert len(ws._images) == 0            # nothing embedded, no crash
    assert any(c.value == "No Photo" for row in ws.iter_rows() for c in row)
